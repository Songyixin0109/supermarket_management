from openpyxl import load_workbook

from django.shortcuts import render, redirect
from django.core.exceptions import ValidationError
from django.utils.safestring import mark_safe
from django import forms
from django.db.models import Prefetch,F
from django.contrib.auth.hashers import make_password
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse, HttpResponse

from OnlineShop import models
from OnlineShop.models import MerchantItems,InventoryItems
from OnlineShop.utils.bootstrap import BootStrapModelForm
from OnlineShop.utils.encrypt import md5
from OnlineShop.utils.pagination import Pagination

class InventoryItemsInfoModelForm(forms.ModelForm):
    class Meta:
        model = models.InventoryItems
        fields = '__all__'

def inventory_info_admin(request):
    search_data = request.GET.get('search_data', '')
    queryset = InventoryItems.objects.select_related('items_name').prefetch_related(
        Prefetch(
            'items_name__merchant_items_set',
            queryset=MerchantItems.objects.select_related('merchant')
        )
    )

    if search_data:
        queryset = queryset.filter(items_name__name__icontains=search_data)

    page_object = Pagination(request, queryset)

    context = {
        'inventory_items': page_object.page_queryset,
        'title': '库存商品',
        'search_data': search_data,
        'page_string': page_object.html(),
    }
    return render(request, 'inventory/inventory_info_admin.html', context)

def inventory_info_user(request):
    search_data = request.GET.get('search_data', '')
    queryset = InventoryItems.objects.select_related('items_name').prefetch_related(
        Prefetch(
            'items_name__merchant_items_set',
            queryset=MerchantItems.objects.select_related('merchant')
        )
    )

    if search_data:
        queryset = queryset.filter(items_name__name__icontains=search_data)

    page_object = Pagination(request, queryset)

    context = {
        'inventory_items': page_object.page_queryset,
        'title': '商品列表',
        'search_data': search_data,
        'page_string': page_object.html(),
    }
    return render(request, 'inventory/inventory_info_user.html', context)
    

def inventory_delete(request, nid):
    models.InventoryItems.objects.get(id=nid).delete()
    return redirect('/inventory/info/admin/')

def inventory_statistics(request):
    title='库存统计'
    if request.method == 'GET':
        return render(request,'inventory/inventory_statistics.html',{'title':title})

def inventory_chart(request):
    items = InventoryItems.objects.all()

    x_axis = [item.name for item in items]  # 获取所有商品名称
    series_list = [
        {
        "name": "销售数量",
        "type": "bar",
        "data": [item.inventory_quantity for item in items]  # 获取所有库存数量
        },
        {
        "name": "库存数量",
        "type": "bar",
        "data": [item.inventory_quantity+20 for item in items]  # 获取所有库存数量
        }
    ]

    result = {
        'status': True,
        'data': {
            'legend': ["销售数量","库存数量"],
            'series_list': series_list,
            'x_axis': x_axis,
        }
    }
    return JsonResponse(result)

def inventory_upload(request):
    file_object = request.FILES.get('excel')
    wb = load_workbook(file_object)
    sheet = wb.worksheets[3]

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        exists = models.InventoryItems.objects.filter(name=row[0].value).exists()
        if not exists:
            models.InventoryItems.objects.create(
                name = row[0].value,
                description = row[1].value,
                catalog = row[2].value,
                ask_price = row[3].value,
                inventory_quantity = row[4].value,
                )
    return redirect('/inventory/info/employee/')