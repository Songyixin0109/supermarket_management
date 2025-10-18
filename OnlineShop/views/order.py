from decimal import Decimal
from math import ceil
from openpyxl import load_workbook

from django.shortcuts import render, redirect
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
from django.db.models import Q
from django.utils import timezone
from django.db import transaction

from OnlineShop import models
from OnlineShop.utils.bootstrap import BootStrapModelForm
from OnlineShop.utils.encrypt import md5
from OnlineShop.utils.pagination import Pagination
from OnlineShop.models import EmployeeInfo

class SellOrderInfoModelForm(BootStrapModelForm):
    class Meta:
        model = models.SellOrders
        fields = '__all__'

# 把 choices 做成字典，方便反向查
STATUS_MAP = {label: val for val, label in models.SellOrders.status_choices}

def sellorder_info_admin(request):
    search_data = request.GET.get('search_data', '').strip()
    queryset = models.SellOrders.objects.select_related('user', 'inventory_items').order_by('updated_at')

    if search_data:
        # 先判断是不是中文状态
        status_num = STATUS_MAP.get(search_data)          # 找不到返回 None
        q = Q(id__icontains=search_data) | Q(user__username__icontains=search_data)
        if status_num is not None:                        # 输入的是“已完成”等
            q |= Q(status=status_num)
        queryset = queryset.filter(q)

    page_object = Pagination(request, queryset)
    context = {
        'sell_orders': page_object.page_queryset,
        'search_data': search_data,
        'page_string': page_object.html(),
        'title': '所有订单',
        'form': SellOrderInfoModelForm(),
    }
    return render(request, 'order/sellorder_info_admin.html', context)

class SellOrderAdminEditForm(BootStrapModelForm):
    """管理员编辑：仅 status 可改"""
    class Meta:
        model = models.SellOrders
        fields = '__all__'       

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        # 除了 status 以外全部禁用
        for name, field in self.fields.items():
            if name != 'status':
                field.disabled = True

@csrf_exempt
def sellorder_admin_edit(request, nid):
    order = models.SellOrders.objects.filter(id=nid).first()
    if not order:
        return JsonResponse({'status': False, 'error': '订单不存在'})
    if request.method == 'GET':
        fields = {}
        fields['id']               = order.id
        fields['用户']             = order.user.username if order.user else '-'
        fields['商品']             = str(order.inventory_items) if order.inventory_items else '-'
        fields['购买数量']          = order.order_quantity
        fields['状态']              = order.status          # 当前值
        fields['状态可选值']        = models.SellOrders.status_choices  # 下拉框用
        return JsonResponse({'status': True, 'fields': fields})
    new_status = int(request.POST.get("status"))
    if new_status == order.status:
        return JsonResponse({"status": False, "error": {"status": ["您没有改变状态！"]}})

    models.SellOrders.objects.filter(id=nid).update(
        status=new_status,
        updated_at=timezone.now()
    )
    return JsonResponse({"status": True, "msg": "状态已更新"})

def sellorder_info_user(request):
    search_data = request.GET.get('search_data', '').strip()
    queryset = models.SellOrders.objects.filter(
        user_id=request.session['info']['id'],
        is_deleted=False
    ).select_related('user', 'inventory_items').order_by('updated_at')
    if search_data:
        queryset = queryset.filter(
            Q(id__icontains=search_data) |
            Q(inventory_items__items_name__name__icontains=search_data)
        )
    page_object = Pagination(request, queryset)
    context = {
        'sell_orders': page_object.page_queryset,
        'search_data': search_data,
        'page_string': page_object.html(),
        'title': '我的订单',
        'form': SellOrderInfoModelForm(),
    }
    return render(request, 'order/sellorder_info_user.html', context)

def sellorder_delete_user(request, nid):
    models.SellOrders.objects.filter(id=nid).update(is_deleted=True)  # 只改标记，不删数据
    return redirect('/sellorder/info/user/')


class SellOrderAddModelForm(BootStrapModelForm):
    class Meta:
        model = models.SellOrders
        exclude = ['id', 'user', 'status', 'order_date', 'updated_at']   
@csrf_exempt
def sellorder_add(request):
    if request.method != 'POST':
        return JsonResponse({'status': False, 'error': '仅支持 POST'})

    form = SellOrderAddModelForm(data=request.POST)
    if not form.is_valid():
        return JsonResponse({'status': False, 'error': form.errors})

    inventory_items_id = form.cleaned_data['inventory_items'].id
    order_quantity = form.cleaned_data['order_quantity']

    # 原子事务：锁库存，防止并发超卖
    with transaction.atomic():
        # 加锁查询库存
        inventory = (
            models.InventoryItems.objects.select_for_update()
            .filter(id=inventory_items_id)
            .first()
        )
        if not inventory:
            return JsonResponse({'status': False, 'error': '商品不存在'})

        if inventory.inventory_quantity < order_quantity:
            return JsonResponse({'status': False, 'error': '库存不足，无法下单'})

        # 扣减库存
        inventory.inventory_quantity -= order_quantity
        inventory.save()

        # 创建订单
        sell_order = form.save(commit=False)
        sell_order.user_id = request.session['info']['id']
        sell_order.status = 1
        sell_order.is_deleted = False
        sell_order.save()

        # 计算总金额
        total = order_quantity * inventory.sell_price

    return JsonResponse({
        'status': True,
        'order_id': sell_order.id,
        'total_amount': float(total)
    })





# ===================入库单====================




class PurchaseOrderInfoModelForm(BootStrapModelForm):
    class Meta:
        model = models.PurchaseOrders
        fields = '__all__'

STATUS_MAP = {label: val for val, label in models.PurchaseOrders.STATUS_CHOICES}

def purchaseorder_info(request):
    search_data = request.GET.get('search_data', '').strip()
    queryset = models.PurchaseOrders.objects.select_related(
        'merchant_items__merchant', 'employee').order_by('updated_at')
    if search_data:
        status_num = STATUS_MAP.get(search_data)        
        # 修复搜索条件：对 ForeignKey 字段使用正确的查找方式
        q = Q(id__icontains=search_data) | \
            Q(merchant_items__merchant_items__name__icontains=search_data) | \
            Q(employee__employee_name__icontains=search_data)
        
        if status_num is not None:                      
            q |= Q(status=status_num)
        queryset = queryset.filter(q)

    page_object = Pagination(request, queryset)
    form = PurchaseOrderInfoModelForm()
    title='所有入库单'
    context = {
        'form': form,
        'purchase_orders': page_object.page_queryset,
        'title':title,
        'search_data':search_data,
        'page_string':page_object.html()}
    return render(request, 'order/purchase_info.html', context)
    
class PurchaseOrderStatusForm(BootStrapModelForm):
    """管理员编辑：仅 status 可改"""
    class Meta:
        model = models.PurchaseOrders
        fields = '__all__'         

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        # 除了 status 以外全部禁用
        for name, field in self.fields.items():
            if name != 'status':
                field.disabled = True

@csrf_exempt
def purchase_admin_edit(request, nid):
    order = (
        models.PurchaseOrders.objects.select_related(
            "merchant_items__merchant_items",
            "merchant_items__merchant",
            "employee",
        )
        .filter(id=nid)
        .first()
    )
    if not order:
        return JsonResponse({"status": False, "error": "入库单不存在"})

    if request.method == "GET":
        fields = {}
        fields["id"] = order.id
        fields["merchant_items"] = (
            order.merchant_items.merchant_items.name
            if order.merchant_items
            else "-"
        )
        fields["merchant_name"] = (
            order.merchant_items.merchant.merchant_name
            if order.merchant_items
            else "-"
        )
        fields["merchant_price"] = (
            str(order.merchant_items.merchant_price)
            if order.merchant_items
            else "-"
        )
        fields["purchase_quantity"] = order.purchase_quantity
        fields["employee"] = (
            order.employee.employee_name if order.employee else "-"
        )
        fields["status"] = order.status
        fields["created_at"] = order.created_at.strftime("%Y-%m-%d %H:%M")
        fields["updated_at"] = order.updated_at.strftime("%Y-%m-%d %H:%M")
        fields["status_choices"] = models.PurchaseOrders.STATUS_CHOICES
        return JsonResponse({"status": True, "fields": fields})

    new_status = int(request.POST.get("status", -1))
    original_status = order.status
    if new_status == original_status:
        return JsonResponse({"status": False, "error": {"status": ["您没有改变状态！"]}})

    # ① 手动刷新 updated_at（auto_now 不会被 update 触发）
    models.PurchaseOrders.objects.filter(id=nid).update(
        status=new_status, updated_at=timezone.now()
    )

    # ② 已入库(2) 时 自动写 / 更新 InventoryItems
    if new_status == 2:
        order.refresh_from_db()
        merchant_items = order.merchant_items
        if merchant_items:
            sell_price = int(ceil(merchant_items.merchant_price * Decimal("1.3")))
            qty = order.purchase_quantity

            inv, created = models.InventoryItems.objects.get_or_create(
                items_name=merchant_items.merchant_items,
                defaults={"sell_price": sell_price, "inventory_quantity": qty},
            )
            if not created:
                inv.inventory_quantity += qty
                inv.sell_price = sell_price
                inv.save()

    return JsonResponse({"status": True, "msg": "状态已更新"})

@csrf_exempt
def purchase_order_upload(request):
    if request.method != 'POST':
        return JsonResponse({'status': False, 'msg': '仅支持 POST'})

    file = request.FILES.get('excel')
    if not file:
        return JsonResponse({'status': False, 'msg': '未选择文件'})

    # 取当前登录员工
    user_id = request.session.get('info', {}).get('id')          # 当前登录用户ID
    emp = EmployeeInfo.objects.filter(id=user_id).first()
    if not emp:
        return JsonResponse({'status': False, 'msg': '未找到当前员工信息，请重新登录'})

    try:
        wb = load_workbook(file)
        sheet = wb.worksheets[0]

        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
            item_name, merchant_name, price, desc, catalog, qty = row[:6]

            if not all([item_name, merchant_name, price, qty]):
                continue            # 必填列有空值则跳过

            item_obj, _ = models.PurchaseItems.objects.get_or_create(
                name=item_name.strip(),
                defaults={'description': desc or '', 'catalog': catalog or ''}
            )

            merchant_obj, _ = models.MerchantInfo.objects.get_or_create(
                merchant_name=merchant_name.strip(),
                defaults={'phone': ''}
            )

            merchant_item_obj, _ = models.MerchantItems.objects.get_or_create(
                merchant=merchant_obj,
                merchant_items=item_obj,
                defaults={'merchant_price': price}
            )

            models.PurchaseOrders.objects.create(
                merchant_items=merchant_item_obj,
                purchase_quantity=qty,
                employee=emp,          
                status=1
            )

        return JsonResponse({'status': True, 'msg': '导入成功'})
    except Exception as e:
        return JsonResponse({'status': False, 'msg': f'导入失败：{str(e)}'})
    

